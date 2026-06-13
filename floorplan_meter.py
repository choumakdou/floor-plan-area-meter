"""
Floor Plan Area Meter
=====================
A desktop tool for measuring room areas on a floor plan image.

Features
--------
- Import a floor plan image (PNG / JPG / BMP / GIF / TIFF).
- Calibrate the image by setting two reference points on the vertical axis
  and two on the horizontal axis, each with a known real-world distance
  (in metres or feet).
- Click on the image to place polygon vertices; close the polygon by
  right-clicking (or pressing Enter). Each polygon gets a random
  semi-transparent colour (alpha = 0.5) and a live area label.
- Toggle the area units live: square points (image px²), square feet,
  or square metres.
- Zoom & pan to navigate large floor plans.
- Save / load measured polygons as JSON so you can re-open a project.

Run:  python floorplan_meter.py
Build with PyInstaller:
    pyinstaller --noconfirm --onefile --windowed \
        --name FloorPlanMeter floorplan_meter.py
"""

from __future__ import annotations

import json
import math
import os
import random
import tkinter as tk
from tkinter import colorchooser, filedialog, messagebox, ttk, simpledialog
from typing import Any, Dict, List, Optional, Tuple

# Build metadata — stamped by the GitHub Actions workflow on every build.
# When running from source these fall back to the dev defaults below.
# --- build metadata ---
__version__ = "2.1.0-measurement"
__build__ = "local"
__built_at__ = ""

# Author / brand credit.  Shown in the window title and Help -> About.
APP_TITLE = "Floor Plan Area Meter"
APP_AUTHOR = ""
APP_COMPANY = "CHFT"

try:
    from PIL import Image, ImageTk, ImageDraw
except ImportError:  # pragma: no cover
    raise SystemExit(
        "Pillow is required.  Install it with:  pip install pillow"
    )

Point = Tuple[float, float]
Polygon = List[Point]


# ---------------------------------------------------------------------------
# v2.1: Measurement sheet — arithmetic L x W breakdown of room areas.
# Mirrors the structure of the user's CHFT Excel "221 1F to print (draft)"
# sheet: each row is a (sub-)rectangle or (sub-)triangle of a room, with
# sub-segment H and V inputs in feet + inches, a shape factor, and an
# optional link to a polygon for cross-check.
# ---------------------------------------------------------------------------
SHAPE_RECT = "rectangle"
SHAPE_TRI  = "triangle"
SHAPE_FACTORS = {SHAPE_RECT: 1.0, SHAPE_TRI: 0.5}


class MeasurementRow:
    """One row in the measurement sheet.

    A row decomposes a room (or a sub-portion of one) into a primitive shape
    whose area is H_feet * V_feet * factor.  Sub-segments let you build a
    composite measurement out of several smaller lengths added together,
    exactly like the user's Excel formulas:
        H = E10 + H10 + K10 + (F10 + I10 + L10) / 12
    """

    __slots__ = (
        "id", "description", "shape", "factor",
        "h_subs",      # 3 (ft, in) pairs
        "v_subs",      # 3 (ft, in) pairs
        "polygon_idx", # optional index into App.polygons, or None
    )

    _next_id = 1

    def __init__(self,
                 description: str = "",
                 shape: str = SHAPE_RECT,
                 h_subs: Optional[List[Tuple[float, float]]] = None,
                 v_subs: Optional[List[Tuple[float, float]]] = None,
                 polygon_idx: Optional[int] = None) -> None:
        self.id = MeasurementRow._next_id
        MeasurementRow._next_id += 1
        self.description = description or f"Row {self.id}"
        self.shape = shape if shape in SHAPE_FACTORS else SHAPE_RECT
        self.factor = SHAPE_FACTORS[self.shape]
        # Pad/truncate sub-segments to exactly 3 pairs of (feet, inches).
        self.h_subs = self._normalise_subs(h_subs)
        self.v_subs = self._normalise_subs(v_subs)
        self.polygon_idx = polygon_idx

    @staticmethod
    def _normalise_subs(subs):
        out = [(0.0, 0.0), (0.0, 0.0), (0.0, 0.0)]
        if subs:
            for i, v in enumerate(subs[:3]):
                if isinstance(v, (list, tuple)) and len(v) == 2:
                    out[i] = (float(v[0]), float(v[1]))
        return out

    # ------------------------------------------------------------------
    def h_feet_total(self) -> float:
        return sum(ft + inch / 12.0 for ft, inch in self.h_subs)

    def v_feet_total(self) -> float:
        return sum(ft + inch / 12.0 for ft, inch in self.v_subs)

    def area_sqft(self) -> float:
        return self.h_feet_total() * self.v_feet_total() * self.factor

    def to_dict(self) -> Dict[str, Any]:
        return {
            "id": self.id,
            "description": self.description,
            "shape": self.shape,
            "h_subs": [list(s) for s in self.h_subs],
            "v_subs": [list(s) for s in self.v_subs],
            "polygon_idx": self.polygon_idx,
        }

    @classmethod
    def from_dict(cls, d: Dict[str, Any]) -> "MeasurementRow":
        row = cls(
            description=d.get("description", ""),
            shape=d.get("shape", SHAPE_RECT),
            h_subs=d.get("h_subs"),
            v_subs=d.get("v_subs"),
            polygon_idx=d.get("polygon_idx"),
        )
        if isinstance(d.get("id"), int):
            row.id = d["id"]
            MeasurementRow._next_id = max(MeasurementRow._next_id, d["id"] + 1)
        return row



# ---------------------------------------------------------------------------
# Geometry helpers
# ---------------------------------------------------------------------------
def polygon_area_image_pts(poly: Polygon) -> float:
    """Shoelace area of a polygon in raw image-pixel² units."""
    if len(poly) < 3:
        return 0.0
    s = 0.0
    n = len(poly)
    for i in range(n):
        x1, y1 = poly[i]
        x2, y2 = poly[(i + 1) % n]
        s += x1 * y2 - x2 * y1
    return abs(s) / 2.0


# ---------------------------------------------------------------------------
# Calibration: convert image-pixel² to a real-world area
# ---------------------------------------------------------------------------
class Calibration:
    """
    Holds two V reference points (p1, p2) with a real distance Vd,
    and two H reference points (p3, p4) with a real distance Hd.
    Provides px² → m² and px² → ft² conversions.
    """

    REAL_UNITS = ("m", "ft")

    def __init__(self) -> None:
        self.p1: Optional[Point] = None
        self.p2: Optional[Point] = None
        self.p3: Optional[Point] = None
        self.p4: Optional[Point] = None
        self.v_real: float = 0.0      # distance for V reference, in chosen unit
        self.h_real: float = 0.0
        self.unit: str = "m"          # 'm' or 'ft'

    # ------------------------------------------------------------------
    @property
    def ready(self) -> bool:
        return all(p is not None for p in (self.p1, self.p2, self.p3, self.p4)) \
            and self.v_real > 0 and self.h_real > 0

    @property
    def px_per_v(self) -> float:
        if not (self.p1 and self.p2):
            return 0.0
        dx = self.p2[0] - self.p1[0]
        dy = self.p2[1] - self.p1[1]
        return math.hypot(dx, dy)

    @property
    def px_per_h(self) -> float:
        if not (self.p3 and self.p4):
            return 0.0
        dx = self.p4[0] - self.p3[0]
        dy = self.p4[1] - self.p3[1]
        return math.hypot(dx, dy)

    def real_per_px_v(self) -> float:
        if self.px_per_v == 0:
            return 0.0
        return self.v_real / self.px_per_v

    def real_per_px_h(self) -> float:
        if self.px_per_h == 0:
            return 0.0
        return self.h_real / self.px_per_h

    # ------------------------------------------------------------------
    def area_real(self, area_px2: float) -> Tuple[float, float]:
        """
        Return (area in m², area in ft²) for a polygon whose image-pixel
        area is `area_px2`.  Combines horizontal & vertical scales so
        that axes with different scales (rare but possible) still work.
        """
        if not self.ready:
            return 0.0, 0.0
        # (m_per_px_h * m_per_px_v) * px²  →  m²
        m_per_px_h = self.real_per_px_h()
        m_per_px_v = self.real_per_px_v()
        m2 = area_px2 * m_per_px_h * m_per_px_v
        ft2 = m2 * 10.7639
        return m2, ft2

    # ------------------------------------------------------------------
    def to_dict(self) -> dict:
        return {
            "p1": self.p1, "p2": self.p2,
            "p3": self.p3, "p4": self.p4,
            "v_real": self.v_real, "h_real": self.h_real,
            "unit": self.unit,
        }

    def from_dict(self, d: dict) -> None:
        self.p1 = tuple(d["p1"]) if d.get("p1") else None
        self.p2 = tuple(d["p2"]) if d.get("p2") else None
        self.p3 = tuple(d["p3"]) if d.get("p3") else None
        self.p4 = tuple(d["p4"]) if d.get("p4") else None
        self.v_real = float(d.get("v_real", 0.0))
        self.h_real = float(d.get("h_real", 0.0))
        self.unit = d.get("unit", "m")


# ---------------------------------------------------------------------------
# Main application
# ---------------------------------------------------------------------------
class App(tk.Tk):
    COLOURS = [
        "#e6194B", "#3cb44b", "#4363d8", "#f58231", "#911eb4",
        "#42d4f4", "#f032e6", "#bfef45", "#fabed4", "#469990",
        "#dcbeff", "#9A6324", "#fffac8", "#800000", "#aaffc3",
        "#808000", "#ffd8b1", "#000075",
    ]

    ZOOM_STEP = 1.25
    MIN_ZOOM = 0.05
    MAX_ZOOM = 20.0

    def __init__(self) -> None:
        super().__init__()
        credit = f"{APP_AUTHOR}, {APP_COMPANY}" if APP_AUTHOR else APP_COMPANY
        self.title(f"{APP_TITLE}  —  {credit}")
        self.geometry("1280x820")
        self.minsize(900, 600)

        # state
        self.image: Optional[Image.Image] = None
        self.image_path: Optional[str] = None
        self.tk_image: Optional[ImageTk.PhotoImage] = None
        self.zoom = 1.0
        self.offset = (0.0, 0.0)        # pan offset in canvas coords
        self._panning = False
        self._pan_start = (0.0, 0.0)

        self.calib = Calibration()
        self.mode = tk.StringVar(value="polygon")     # 'polygon' | 'calibV' | 'calibH' | 'edit'
        self.input_mode = tk.StringVar(value="draw")  # 'draw' | 'pan'  (interaction mode)
        self.unit_var = tk.StringVar(value="m²")
        self.show_labels = tk.BooleanVar(value=True)
        self.show_calib = tk.BooleanVar(value=True)

        # Edit-mode state
        self.selected_polygon: Optional[int] = None
        self.selected_vertex: Optional[int] = None
        self._dragging_vertex = False
        self._hover_edge: Optional[Tuple[int, int]] = None  # (poly_idx, edge_idx)

        # v2.1: measurement sheet (arithmetic L x W breakdown)
        self.measurement_rows: List["MeasurementRow"] = []

        # polygons: list of dicts {poly, color, name}
        self.polygons: List[dict] = []
        self.current_poly: Polygon = []
        self._hover_point: Optional[Point] = None

        # canvas-side caches (always use image-space coords)
        self._display_image: Optional[Image.Image] = None

        self._build_ui()
        self._bind_events()
        self._update_status()

    # ------------------------------------------------------------------
    # UI
    # ------------------------------------------------------------------
    def _build_ui(self) -> None:
        # Toolbar
        tb = ttk.Frame(self, padding=4)
        tb.pack(side=tk.TOP, fill=tk.X)

        ttk.Button(tb, text="Open Image…", command=self.open_image).pack(side=tk.LEFT, padx=2)
        ttk.Button(tb, text="Save Project…", command=self.save_project).pack(side=tk.LEFT, padx=2)
        ttk.Button(tb, text="Load Project…", command=self.load_project).pack(side=tk.LEFT, padx=2)
        ttk.Separator(tb, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=6)

        ttk.Label(tb, text="Mode:").pack(side=tk.LEFT)
        ttk.Radiobutton(tb, text="Polygon", variable=self.mode,
                        value="polygon", command=self._on_mode_change).pack(side=tk.LEFT)
        ttk.Radiobutton(tb, text="Edit", variable=self.mode,
                        value="edit", command=self._on_mode_change).pack(side=tk.LEFT)
        ttk.Radiobutton(tb, text="Calibrate V", variable=self.mode,
                        value="calibV", command=self._on_mode_change).pack(side=tk.LEFT)
        ttk.Radiobutton(tb, text="Calibrate H", variable=self.mode,
                        value="calibH", command=self._on_mode_change).pack(side=tk.LEFT)
        ttk.Separator(tb, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=6)

        # Draw / Pan toggle (Adobe-style interaction switch)
        self.input_mode_btn = ttk.Checkbutton(
            tb, text="Pan mode", variable=self.input_mode,
            onvalue="pan", offvalue="draw",
            command=self._on_input_mode_change)
        self.input_mode_btn.pack(side=tk.LEFT, padx=2)
        ttk.Separator(tb, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=6)

        ttk.Label(tb, text="Units:").pack(side=tk.LEFT)
        unit_combo = ttk.Combobox(tb, textvariable=self.unit_var, width=8,
                                   values=("px²", "m²", "ft²"),
                                   state="readonly")
        unit_combo.pack(side=tk.LEFT)
        unit_combo.bind("<<ComboboxSelected>>", lambda e: self.redraw())

        ttk.Checkbutton(tb, text="Labels", variable=self.show_labels,
                        command=self.redraw).pack(side=tk.LEFT, padx=4)
        ttk.Checkbutton(tb, text="Calibration", variable=self.show_calib,
                        command=self.redraw).pack(side=tk.LEFT, padx=4)
        ttk.Separator(tb, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=6)

        ttk.Button(tb, text="Undo Point", command=self.undo_point).pack(side=tk.LEFT, padx=2)
        ttk.Button(tb, text="Close Polygon", command=self.close_polygon).pack(side=tk.LEFT, padx=2)
        ttk.Button(tb, text="Clear Current", command=self.clear_current).pack(side=tk.LEFT, padx=2)
        ttk.Button(tb, text="Delete Last Polygon", command=self.delete_last_polygon).pack(side=tk.LEFT, padx=2)
        ttk.Button(tb, text="Clear All Polygons", command=self.clear_all_polygons).pack(side=tk.LEFT, padx=2)
        ttk.Separator(tb, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=6)

        # v2.1: measurement sheet actions
        ttk.Button(tb, text="+ Row", command=self._add_measurement_row).pack(side=tk.LEFT, padx=2)
        ttk.Button(tb, text="Export xlsx", command=self._export_measurement_sheet_xlsx).pack(side=tk.LEFT, padx=2)
        ttk.Separator(tb, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=6)

        ttk.Button(tb, text="Zoom +", command=lambda: self.zoom_by(self.ZOOM_STEP)).pack(side=tk.LEFT, padx=2)
        ttk.Button(tb, text="Zoom −", command=lambda: self.zoom_by(1 / self.ZOOM_STEP)).pack(side=tk.LEFT, padx=2)
        ttk.Button(tb, text="Fit", command=self.fit_to_window).pack(side=tk.LEFT, padx=2)
        ttk.Button(tb, text="100%", command=lambda: self.set_zoom(1.0)).pack(side=tk.LEFT, padx=2)

        # Persistent stringvars that need to be created early
        self.unit_label_var = tk.StringVar(value="m")
        self.calib_unit = tk.StringVar(value="m")

        # App-level menu (Help → About shows build / version info)
        menubar = tk.Menu(self)
        helpmenu = tk.Menu(menubar, tearoff=0)
        helpmenu.add_command(label="About", command=self._show_about)
        menubar.add_cascade(label="Help", menu=helpmenu)
        self.config(menu=menubar)

        # Main paned layout: canvas on the left, controls on the right
        main = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        main.pack(fill=tk.BOTH, expand=True)

        # Canvas
        self.canvas = tk.Canvas(main, background="#202225", highlightthickness=0)
        main.add(self.canvas, weight=5)

        # Side panel
        side = ttk.Frame(main, padding=8)
        main.add(side, weight=1)

        # ----- calibration panel -----
        cal = ttk.LabelFrame(side, text="Calibration", padding=8)
        cal.pack(fill=tk.X, pady=4)

        ttk.Label(cal, text=(
            "1. Pick mode  •  V or H\n"
            "2. Click two points on the image\n"
            "3. Enter the real distance\n\n"
            "Provide both V and H references."
        ), justify=tk.LEFT).pack(anchor=tk.W)

        ttk.Separator(cal, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=6)

        # V inputs
        ttk.Label(cal, text="Vertical reference (P1 → P2):").pack(anchor=tk.W)
        vrow = ttk.Frame(cal); vrow.pack(fill=tk.X, pady=2)
        ttk.Label(vrow, text="x₁").pack(side=tk.LEFT)
        self.v_x1 = tk.StringVar(); ttk.Entry(vrow, textvariable=self.v_x1, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Label(vrow, text="y₁").pack(side=tk.LEFT)
        self.v_y1 = tk.StringVar(); ttk.Entry(vrow, textvariable=self.v_y1, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Label(vrow, text="x₂").pack(side=tk.LEFT)
        self.v_x2 = tk.StringVar(); ttk.Entry(vrow, textvariable=self.v_x2, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Label(vrow, text="y₂").pack(side=tk.LEFT)
        self.v_y2 = tk.StringVar(); ttk.Entry(vrow, textvariable=self.v_y2, width=8).pack(side=tk.LEFT, padx=2)
        vrow2 = ttk.Frame(cal); vrow2.pack(fill=tk.X, pady=2)
        ttk.Label(vrow2, text="Distance:").pack(side=tk.LEFT)
        self.v_dist = tk.StringVar(value="0")
        ttk.Entry(vrow2, textvariable=self.v_dist, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Label(vrow2, textvariable=self.unit_label_var).pack(side=tk.LEFT)
        ttk.Button(vrow2, text="Set V", command=self.set_v_from_inputs).pack(side=tk.LEFT, padx=4)


        # H inputs
        ttk.Label(cal, text="Horizontal reference (P3 → P4):").pack(anchor=tk.W, pady=(8, 0))
        hrow = ttk.Frame(cal); hrow.pack(fill=tk.X, pady=2)
        ttk.Label(hrow, text="x₁").pack(side=tk.LEFT)
        self.h_x1 = tk.StringVar(); ttk.Entry(hrow, textvariable=self.h_x1, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Label(hrow, text="y₁").pack(side=tk.LEFT)
        self.h_y1 = tk.StringVar(); ttk.Entry(hrow, textvariable=self.h_y1, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Label(hrow, text="x₂").pack(side=tk.LEFT)
        self.h_x2 = tk.StringVar(); ttk.Entry(hrow, textvariable=self.h_x2, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Label(hrow, text="y₂").pack(side=tk.LEFT)
        self.h_y2 = tk.StringVar(); ttk.Entry(hrow, textvariable=self.h_y2, width=8).pack(side=tk.LEFT, padx=2)
        hrow2 = ttk.Frame(cal); hrow2.pack(fill=tk.X, pady=2)
        ttk.Label(hrow2, text="Distance:").pack(side=tk.LEFT)
        self.h_dist = tk.StringVar(value="0")
        ttk.Entry(hrow2, textvariable=self.h_dist, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Label(hrow2, textvariable=self.unit_label_var).pack(side=tk.LEFT)
        ttk.Button(hrow2, text="Set H", command=self.set_h_from_inputs).pack(side=tk.LEFT, padx=4)

        ttk.Separator(cal, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=6)
        unit_row = ttk.Frame(cal); unit_row.pack(fill=tk.X)
        ttk.Label(unit_row, text="Real unit:").pack(side=tk.LEFT)
        unit_cb = ttk.Combobox(unit_row, textvariable=self.calib_unit,
                                values=Calibration.REAL_UNITS,
                                state="readonly", width=6)
        unit_cb.pack(side=tk.LEFT, padx=4)
        unit_cb.bind("<<ComboboxSelected>>", lambda e: self._refresh_unit_label())
        self._refresh_unit_label()

        self.calib_status = ttk.Label(cal, text="Not calibrated", foreground="#aa2222")
        self.calib_status.pack(anchor=tk.W, pady=(4, 0))

        # ----- side-panel notebook: Polygons / Measurement Sheet -----
        side_nb = ttk.Notebook(side)
        side_nb.pack(fill=tk.BOTH, expand=True, pady=4)

        # --- Polygons tab (was the old listbox) ---
        plf = ttk.Frame(side_nb, padding=8)
        side_nb.add(plf, text="Polygons")
        self.poly_list = tk.Listbox(plf, height=10, exportselection=False)
        self.poly_list.pack(fill=tk.BOTH, expand=True)
        self.poly_list.bind("<<ListboxSelect>>", self._on_poly_list_select)

        # --- Measurement Sheet tab (v2.1) ---
        self._build_measurement_sheet_tab(side_nb)

        # ----- legend / tips -----
        tips = ttk.LabelFrame(side, text="Tips", padding=8)
        tips.pack(fill=tk.X, pady=4)
        ttk.Label(tips, justify=tk.LEFT, text=(
            "• Draw mode: left-click = place point\n"
            "• Pan mode: left-drag = pan the canvas\n"
            "• Mouse-wheel: zoom (Ctrl+wheel for finer)\n"
            "• Middle-drag or Space+drag: pan (always)\n"
            "• Arrow keys: nudge view\n"
            "• Right-click: close polygon / delete vertex (Edit)\n"
            "• Enter: close polygon\n"
            "• Esc: cancel current polygon\n"
            "\n"
            "Edit mode:\n"
            "• Click a vertex to select (drag to move)\n"
            "• Double-click an edge to insert a vertex\n"
            "• Right-click a vertex to delete it (>=4 remain)"
        )).pack(anchor=tk.W)

        # Status bar
        self.status = tk.StringVar(value="Open an image to start.")
        sb = ttk.Label(self, textvariable=self.status, anchor=tk.W,
                       relief=tk.SUNKEN, padding=4)
        sb.pack(side=tk.BOTTOM, fill=tk.X)

    def _refresh_unit_label(self) -> None:
        self.unit_label_var.set(self.calib_unit.get())

    # ------------------------------------------------------------------
    # Events
    # ------------------------------------------------------------------
    def _bind_events(self) -> None:
        self.canvas.bind("<Button-1>", self.on_left_click)
        self.canvas.bind("<Button-3>", self.on_right_click)
        self.canvas.bind("<Button-2>", self.on_pan_start)
        self.canvas.bind("<B2-Motion>", self.on_pan_move)
        self.canvas.bind("<ButtonRelease-2>", self.on_pan_end)
        self.canvas.bind("<Motion>", self.on_motion)
        self.canvas.bind("<ButtonRelease-1>", self.on_left_release)
        self.canvas.bind("<Double-Button-1>", self.on_double_click)
        self.canvas.bind("<MouseWheel>", self.on_mousewheel)        # win/mac
        self.canvas.bind("<Button-4>", lambda e: self.zoom_by(self.ZOOM_STEP))   # linux
        self.canvas.bind("<Button-5>", lambda e: self.zoom_by(1 / self.ZOOM_STEP))
        self.canvas.bind("<Configure>", lambda e: self._schedule_redraw())
        self.bind("<Return>", lambda e: self.close_polygon())
        self.bind("<Escape>", lambda e: self.clear_current())
        self.bind("<space>", self._on_space_press)
        self.bind("<KeyRelease-space>", self._on_space_release)
        for k in ("Left", "Right", "Up", "Down"):
            self.bind(f"<{k}>", self._on_arrow_pan)

    def _on_space_press(self, event):
        self.canvas.config(cursor="fleur")

    def _on_space_release(self, event):
        self.canvas.config(cursor="")
        self._panning = False

    def _on_arrow_pan(self, event):
        if self.image is None:
            return
        step = 40
        dx, dy = 0, 0
        if event.keysym == "Left":  dx =  step
        if event.keysym == "Right": dx = -step
        if event.keysym == "Up":    dy =  step
        if event.keysym == "Down":  dy = -step
        ox, oy = self.offset
        self.offset = (ox + dx, oy + dy)
        self._schedule_redraw()

    # ------------------------------------------------------------------
    # Image loading
    # ------------------------------------------------------------------
    def open_image(self) -> None:
        path = filedialog.askopenfilename(
            title="Open floor plan",
            filetypes=[("Images", "*.png *.jpg *.jpeg *.bmp *.tif *.tiff *.gif"),
                       ("All files", "*.*")],
        )
        if not path:
            return
        try:
            img = Image.open(path)
            img.load()
        except Exception as exc:
            messagebox.showerror("Open image", f"Failed to open image:\n{exc}")
            return
        self.image = img
        self.image_path = path
        self.calib = Calibration()
        self.polygons.clear()
        self.current_poly.clear()
        self.poly_list.delete(0, tk.END)
        self.zoom = 1.0
        self.fit_to_window()
        self._update_status()

    # ------------------------------------------------------------------
    # Coordinate transforms
    # ------------------------------------------------------------------
    def canvas_to_image(self, cx: float, cy: float) -> Point:
        ox, oy = self.offset
        return ((cx - ox) / self.zoom, (cy - oy) / self.zoom)

    def image_to_canvas(self, ix: float, iy: float) -> Point:
        ox, oy = self.offset
        return (ix * self.zoom + ox, iy * self.zoom + oy)

    # ------------------------------------------------------------------
    # Zoom / pan
    # ------------------------------------------------------------------
    def zoom_by(self, factor: float, center: Optional[Point] = None) -> None:
        if self.image is None:
            return
        new_zoom = max(self.MIN_ZOOM, min(self.MAX_ZOOM, self.zoom * factor))
        if center is None:
            cw = self.canvas.winfo_width() / 2
            ch = self.canvas.winfo_height() / 2
            center = (cw, ch)
        cx, cy = center
        # keep the image point under the cursor stationary
        ix, iy = self.canvas_to_image(cx, cy)
        self.zoom = new_zoom
        ox, oy = self.offset
        self.offset = (cx - ix * self.zoom, cy - iy * self.zoom)
        self._schedule_redraw()

    def set_zoom(self, z: float) -> None:
        self.zoom_by(z / self.zoom)

    def fit_to_window(self) -> None:
        if self.image is None:
            return
        cw = max(self.canvas.winfo_width(), 1)
        ch = max(self.canvas.winfo_height(), 1)
        iw, ih = self.image.size
        z = min(cw / iw, ch / ih) * 0.98
        self.zoom = max(self.MIN_ZOOM, min(self.MAX_ZOOM, z))
        self.offset = ((cw - iw * self.zoom) / 2, (ch - ih * self.zoom) / 2)
        self._schedule_redraw()

    def on_mousewheel(self, event):
        if self.image is None:
            return
        delta = event.delta
        if delta == 0:
            return
        if event.state & 0x0004:  # ctrl
            delta = max(-1, min(1, delta)) * 1
        factor = self.ZOOM_STEP if delta > 0 else 1 / self.ZOOM_STEP
        self.zoom_by(factor, (event.x, event.y))

    def on_pan_start(self, event):
        self._panning = True
        self._pan_start = (event.x, event.y)
        self.canvas.config(cursor="fleur")

    def on_pan_move(self, event):
        if not self._panning:
            return
        dx = event.x - self._pan_start[0]
        dy = event.y - self._pan_start[1]
        self._pan_start = (event.x, event.y)
        ox, oy = self.offset
        self.offset = (ox + dx, oy + dy)
        self._schedule_redraw()

    def on_pan_end(self, event):
        self._panning = False
        self.canvas.config(cursor="")

    # ------------------------------------------------------------------
    # Click handlers
    # ------------------------------------------------------------------
    def on_left_click(self, event):
        if self.image is None:
            messagebox.showinfo("No image", "Open a floor plan image first.")
            return
        ix, iy = self.canvas_to_image(event.x, event.y)
        mode = self.mode.get()

        # In pan mode, a left click starts a pan (acts like middle-drag).
        if self.input_mode.get() == "pan":
            self._panning = True
            self._pan_start = (event.x, event.y)
            self._pan_origin = self.offset
            return

        if mode == "calibV":
            if self.calib.p1 is None:
                self.calib.p1 = (ix, iy)
            elif self.calib.p2 is None:
                self.calib.p2 = (ix, iy)
            self._refresh_calib_entries()
        elif mode == "calibH":
            if self.calib.p3 is None:
                self.calib.p3 = (ix, iy)
            elif self.calib.p4 is None:
                self.calib.p4 = (ix, iy)
            self._refresh_calib_entries()
        elif mode == "edit":
            hit = self._hit_test_polygon(ix, iy, tol_px=10.0)
            if hit is None:
                self.selected_polygon = None
                self.selected_vertex = None
            else:
                pi, vi = hit
                self.selected_polygon = pi
                self.selected_vertex = vi
                self._dragging_vertex = True
                if 0 <= pi < self.poly_list.size():
                    self.poly_list.selection_clear(0, tk.END)
                    self.poly_list.selection_set(pi)
        else:  # polygon
            self.current_poly.append((ix, iy))
        self._schedule_redraw()

    def on_left_release(self, event):
        self._dragging_vertex = False
        self._panning = False

    def on_double_click(self, event):
        if self.image is None or self.mode.get() != "edit":
            return
        ix, iy = self.canvas_to_image(event.x, event.y)
        edge = self._hit_test_edge(ix, iy, tol_px=10.0)
        if edge is None:
            return
        pi, ei = edge
        poly = self.polygons[pi]["poly"]
        new_vi = (ei + 1) % len(poly)
        poly.insert(new_vi, (ix, iy))
        self.selected_polygon = pi
        self.selected_vertex = new_vi
        self._schedule_redraw()

    def on_right_click(self, event):
        if self.mode.get() == "edit" and self.selected_polygon is not None and self.selected_vertex is not None:
            poly = self.polygons[self.selected_polygon]["poly"]
            if len(poly) > 3:
                del poly[self.selected_vertex]
                self.selected_vertex = None
                self._schedule_redraw()
                return
        self.close_polygon()

    def on_motion(self, event):
        if self.image is None:
            return
        self._hover_point = self.canvas_to_image(event.x, event.y)

        # Live pan with left button in pan mode
        if self._panning and self.input_mode.get() == "pan" and getattr(self, "_pan_origin", None) is not None:
            dx = event.x - self._pan_start[0]
            dy = event.y - self._pan_start[1]
            ox0, oy0 = self._pan_origin
            self.offset = (ox0 + dx, oy0 + dy)
            self._schedule_redraw()
            return

        # Live vertex drag in edit mode
        if self._dragging_vertex and self.mode.get() == "edit" \
                and self.selected_polygon is not None and self.selected_vertex is not None:
            ix, iy = self.canvas_to_image(event.x, event.y)
            poly = self.polygons[self.selected_polygon]["poly"]
            if 0 <= self.selected_vertex < len(poly):
                poly[self.selected_vertex] = (ix, iy)
                self._schedule_redraw()
            return

        # Edge hover (for insert-vertex hint)
        if self.mode.get() == "edit":
            ix, iy = self.canvas_to_image(event.x, event.y)
            self._hover_edge = self._hit_test_edge(ix, iy, tol_px=8.0)
        else:
            self._hover_edge = None

        if self.current_poly and self.mode.get() == "polygon":
            self._schedule_redraw()
        else:
            self._update_coord_status(event.x, event.y)

    def _update_coord_status(self, cx: int, cy: int) -> None:
        if self.image is None:
            return
        ix, iy = self.canvas_to_image(cx, cy)
        self.status.set(
            f"cursor: ({ix:.1f}, {iy:.1f}) px   |   zoom: {self.zoom*100:.0f}%"
        )

    # ------------------------------------------------------------------
    # Calibration helpers
    # ------------------------------------------------------------------
    def _refresh_calib_entries(self) -> None:
        if self.calib.p1:
            self.v_x1.set(f"{self.calib.p1[0]:.1f}"); self.v_y1.set(f"{self.calib.p1[1]:.1f}")
        if self.calib.p2:
            self.v_x2.set(f"{self.calib.p2[0]:.1f}"); self.v_y2.set(f"{self.calib.p2[1]:.1f}")
        if self.calib.p3:
            self.h_x1.set(f"{self.calib.p3[0]:.1f}"); self.h_y1.set(f"{self.calib.p3[1]:.1f}")
        if self.calib.p4:
            self.h_x2.set(f"{self.calib.p4[0]:.1f}"); self.h_y2.set(f"{self.calib.p4[1]:.1f}")
        self._refresh_calib_status()

    def _parse_float(self, s: str) -> Optional[float]:
        try:
            return float(s)
        except (ValueError, TypeError):
            return None

    def set_v_from_inputs(self) -> None:
        d = self._parse_float(self.v_dist.get())
        if d is None or d <= 0:
            messagebox.showerror("Calibration", "Enter a positive V distance.")
            return
        self.calib.v_real = d
        self.calib.unit = self.calib_unit.get()
        self._refresh_calib_status()

    def set_h_from_inputs(self) -> None:
        d = self._parse_float(self.h_dist.get())
        if d is None or d <= 0:
            messagebox.showerror("Calibration", "Enter a positive H distance.")
            return
        self.calib.h_real = d
        self.calib.unit = self.calib_unit.get()
        self._refresh_calib_status()

    def _refresh_calib_status(self) -> None:
        if self.calib.ready:
            self.calib_status.config(
                text=f"Calibrated: 1 {self.calib.unit} = "
                     f"{self.calib.px_per_v/self.calib.v_real:.3f} px (V), "
                     f"{self.calib.px_per_h/self.calib.h_real:.3f} px (H)",
                foreground="#1a7f37")
        else:
            self.calib_status.config(text="Not calibrated", foreground="#aa2222")

    # ------------------------------------------------------------------
    # Polygon management
    # ------------------------------------------------------------------
    def close_polygon(self) -> None:
        if len(self.current_poly) < 3:
            self.status.set("Need at least 3 points to close a polygon.")
            return
        colour = random.choice(self.COLOURS)
        name = f"Room {len(self.polygons) + 1}"
        self.polygons.append({"poly": list(self.current_poly), "color": colour, "name": name})
        self.poly_list.insert(tk.END, name)
        self.current_poly.clear()
        self._schedule_redraw()

    def undo_point(self) -> None:
        if self.current_poly:
            self.current_poly.pop()
            self._schedule_redraw()

    def clear_current(self) -> None:
        self.current_poly.clear()
        self._schedule_redraw()

    def delete_last_polygon(self) -> None:
        if not self.polygons:
            return
        self.polygons.pop()
        if self.poly_list.size() > 0:
            self.poly_list.delete(tk.END)
        if self.selected_polygon is not None and self.selected_polygon >= len(self.polygons):
            self.selected_polygon = None
            self.selected_vertex = None
        self._schedule_redraw()

    def clear_all_polygons(self) -> None:
        if not self.polygons:
            return
        if not messagebox.askyesno("Clear all", "Delete every polygon?"):
            return
        self.polygons.clear()
        self.current_poly.clear()
        self.poly_list.delete(0, tk.END)
        self.selected_polygon = None
        self.selected_vertex = None
        self._schedule_redraw()

    # ------------------------------------------------------------------
    # v2.1: Measurement sheet (Treeview + Excel export)
    # ------------------------------------------------------------------
    # Columns we expose in the on-screen table.  Sub-segments (3 each) are
    # shown as ft and in side-by-side, matching the user's existing CHFT
    # Excel format: H = E + H + K + (F + I + L) / 12.
    SHEET_COLS = (
        "#", "Description", "Shape",
        "H ft1", "H in1", "H ft2", "H in2", "H ft3", "H in3", "H total (ft)",
        "V ft1", "V in1", "V ft2", "V in2", "V ft3", "V in3", "V total (ft)",
        "Factor", "Area (sqft)", "Polygon", "Delta %",
    )
    SHEET_COL_WIDTHS = {
        "#": 32, "Description": 140, "Shape": 70,
        "H ft1": 38, "H in1": 38, "H ft2": 38, "H in2": 38, "H ft3": 38, "H in3": 38,
        "H total (ft)": 70,
        "V ft1": 38, "V in1": 38, "V ft2": 38, "V in2": 38, "V ft3": 38, "V in3": 38,
        "V total (ft)": 70,
        "Factor": 50, "Area (sqft)": 80,
        "Polygon": 60, "Delta %": 60,
    }

    def _build_measurement_sheet_tab(self, notebook: ttk.Notebook) -> None:
        frame = ttk.Frame(notebook, padding=4)
        notebook.add(frame, text="Measurement Sheet")

        hint = ttk.Label(
            frame, justify=tk.LEFT,
            text=("Arithmetic L x W breakdown of room areas, in feet + inches.\n"
                  "Double-click any editable cell to change it.  Each row's H/V\n"
                  "is the sum of three sub-segments, just like the CHFT Excel\n"
                  "worksheet.  'Polygon' links the row to a traced polygon for\n"
                  "automatic cross-check (delta %)."),
        )
        hint.pack(anchor=tk.W)

        # Treeview
        tree_wrap = ttk.Frame(frame)
        tree_wrap.pack(fill=tk.BOTH, expand=True, pady=(4, 0))
        self.sheet_tree = ttk.Treeview(
            tree_wrap, columns=self.SHEET_COLS, show="headings", height=10)
        for c in self.SHEET_COLS:
            self.sheet_tree.heading(c, text=c)
            self.sheet_tree.column(c, width=self.SHEET_COL_WIDTHS.get(c, 60),
                                   anchor=tk.CENTER, stretch=False)
        ysb = ttk.Scrollbar(tree_wrap, orient=tk.VERTICAL,
                            command=self.sheet_tree.yview)
        xsb = ttk.Scrollbar(tree_wrap, orient=tk.HORIZONTAL,
                            command=self.sheet_tree.xview)
        self.sheet_tree.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        self.sheet_tree.grid(row=0, column=0, sticky="nsew")
        ysb.grid(row=0, column=1, sticky="ns")
        xsb.grid(row=1, column=0, sticky="ew")
        tree_wrap.rowconfigure(0, weight=1)
        tree_wrap.columnconfigure(0, weight=1)
        self.sheet_tree.bind("<Double-1>", self._on_sheet_double_click)

        # Totals line + buttons
        bottom = ttk.Frame(frame)
        bottom.pack(fill=tk.X, pady=(4, 0))
        ttk.Button(bottom, text="+ Row", command=self._add_measurement_row).pack(side=tk.LEFT)
        ttk.Button(bottom, text="Delete row", command=self._delete_measurement_row).pack(side=tk.LEFT, padx=4)
        ttk.Button(bottom, text="Export xlsx", command=self._export_measurement_sheet_xlsx).pack(side=tk.LEFT)
        ttk.Button(bottom, text="Clear sheet", command=self._clear_measurement_sheet).pack(side=tk.LEFT, padx=4)
        self.sheet_total_label = ttk.Label(bottom, text="", font=("Segoe UI", 10, "bold"))
        self.sheet_total_label.pack(side=tk.RIGHT)

    def _refresh_measurement_sheet(self) -> None:
        if not hasattr(self, "sheet_tree"):
            return
        for iid in self.sheet_tree.get_children():
            self.sheet_tree.delete(iid)
        total_sqft = 0.0
        for idx, row in enumerate(self.measurement_rows, start=1):
            h_tot = row.h_feet_total()
            v_tot = row.v_feet_total()
            area = row.area_sqft()
            total_sqft += area
            poly_label = ""
            delta_pct = ""
            if row.polygon_idx is not None and 0 <= row.polygon_idx < len(self.polygons):
                poly = self.polygons[row.polygon_idx]
                poly_area_px2 = polygon_area_image_pts(poly["poly"])
                m2, _ = self.calib.area_real(poly_area_px2)
                poly_sqft = m2 * 10.7639
                poly_label = f"#{row.polygon_idx + 1}"
                if area > 1e-6:
                    delta_pct = f"{(poly_sqft - area) / area * 100:+.2f}%"
                else:
                    delta_pct = "\u2014"
            values = [
                idx, row.description, row.shape,
                row.h_subs[0][0], row.h_subs[0][1],
                row.h_subs[1][0], row.h_subs[1][1],
                row.h_subs[2][0], row.h_subs[2][1],
                f"{h_tot:.3f}",
                row.v_subs[0][0], row.v_subs[0][1],
                row.v_subs[1][0], row.v_subs[1][1],
                row.v_subs[2][0], row.v_subs[2][1],
                f"{v_tot:.3f}",
                f"{row.factor:g}", f"{area:,.3f}",
                poly_label, delta_pct,
            ]
            self.sheet_tree.insert("", tk.END, iid=str(row.id), values=values)
        total_m2 = total_sqft / 10.7639
        self.sheet_total_label.config(
            text=f"Total: {total_sqft:,.2f} sqft  /  {total_m2:,.2f} m\u00b2")

    def _add_measurement_row(self) -> None:
        link = (len(self.polygons) - 1) if len(self.polygons) == 1 else None
        row = MeasurementRow(polygon_idx=link)
        self.measurement_rows.append(row)
        self._refresh_measurement_sheet()

    def _delete_measurement_row(self) -> None:
        sel = self.sheet_tree.selection() if hasattr(self, "sheet_tree") else ()
        if not sel:
            messagebox.showinfo("Delete row", "Select a row in the measurement sheet first.")
            return
        del_id = int(sel[0])
        self.measurement_rows = [r for r in self.measurement_rows if r.id != del_id]
        self._refresh_measurement_sheet()

    def _clear_measurement_sheet(self) -> None:
        if not self.measurement_rows:
            return
        if not messagebox.askyesno("Clear sheet", "Delete every measurement row?"):
            return
        self.measurement_rows.clear()
        self._refresh_measurement_sheet()

    def _on_sheet_double_click(self, event) -> None:
        region = self.sheet_tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        row_id = self.sheet_tree.identify_row(event.y)
        col_id = self.sheet_tree.identify_column(event.x)
        if not row_id or not col_id:
            return
        row = next((r for r in self.measurement_rows if str(r.id) == row_id), None)
        if row is None:
            return
        col_idx = int(col_id.lstrip("#")) - 1
        col_name = self.SHEET_COLS[col_idx]
        self._edit_sheet_cell(row, col_name)

    def _edit_sheet_cell(self, row, col_name: str) -> None:
        if col_name == "Description":
            new = simpledialog.askstring("Description", "Description:", initialvalue=row.description, parent=self)
            if new is not None:
                row.description = new
        elif col_name == "Shape":
            new = simpledialog.askstring(
                "Shape", "Shape (rectangle / triangle):",
                initialvalue=row.shape, parent=self)
            if new and new in SHAPE_FACTORS:
                row.shape = new
                row.factor = SHAPE_FACTORS[new]
        elif col_name.startswith("H ft"):
            i = int(col_name.split()[2]) - 1
            ft, inch = row.h_subs[i]
            new = simpledialog.askfloat("H feet", f"H sub {i+1} - feet:", initialvalue=ft, parent=self, minvalue=0)
            if new is not None:
                row.h_subs[i] = (new, inch)
        elif col_name.startswith("H in"):
            i = int(col_name.split()[2]) - 1
            ft, inch = row.h_subs[i]
            new = simpledialog.askfloat("H inches", f"H sub {i+1} - inches:", initialvalue=inch, parent=self, minvalue=0)
            if new is not None:
                row.h_subs[i] = (ft, new)
        elif col_name.startswith("V ft"):
            i = int(col_name.split()[2]) - 1
            ft, inch = row.v_subs[i]
            new = simpledialog.askfloat("V feet", f"V sub {i+1} - feet:", initialvalue=ft, parent=self, minvalue=0)
            if new is not None:
                row.v_subs[i] = (new, inch)
        elif col_name.startswith("V in"):
            i = int(col_name.split()[2]) - 1
            ft, inch = row.v_subs[i]
            new = simpledialog.askfloat("V inches", f"V sub {i+1} - inches:", initialvalue=inch, parent=self, minvalue=0)
            if new is not None:
                row.v_subs[i] = (ft, new)
        elif col_name == "Polygon":
            current = "" if row.polygon_idx is None else str(row.polygon_idx + 1)
            new = simpledialog.askstring(
                "Polygon link",
                f"Link to polygon number (1..{len(self.polygons)}) or blank to unlink:",
                initialvalue=current, parent=self)
            if new is None:
                return
            new = new.strip()
            if not new:
                row.polygon_idx = None
            else:
                try:
                    n = int(new)
                    if 1 <= n <= len(self.polygons):
                        row.polygon_idx = n - 1
                    else:
                        messagebox.showerror("Polygon link", f"Polygon must be 1..{len(self.polygons)}")
                except ValueError:
                    messagebox.showerror("Polygon link", "Enter a number, or leave blank.")
        else:
            return
        self._refresh_measurement_sheet()

    def _export_measurement_sheet_xlsx(self) -> None:
        if not self.measurement_rows:
            messagebox.showinfo("Export", "The measurement sheet is empty.")
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            messagebox.showerror(
                "Export",
                "openpyxl is required for .xlsx export.\n\nInstall it with:\n  pip install openpyxl")
            return
        path = filedialog.asksaveasfilename(
            title="Export measurement sheet",
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
            initialfile=(os.path.splitext(os.path.basename(self.image_path or "plan"))[0] or "plan") + "-measurements.xlsx",
        )
        if not path:
            return
        wb = openpyxl.Workbook()
        bold = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="DDDDDD")
        thin = Side(border_style="thin", color="999999")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Sheet 1: Measurement rows
        ws = wb.active
        ws.title = "Measurements"
        headers = [
            "#", "Description", "Shape",
            "H ft1", "H in1", "H ft2", "H in2", "H ft3", "H in3", "H total (ft)",
            "V ft1", "V in1", "V ft2", "V in2", "V ft3", "V in3", "V total (ft)",
            "Factor", "Area (sqft)", "Area (m\u00b2)", "Polygon", "Delta %",
        ]
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = bold
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")
            c.border = border
        total_sqft = 0.0
        for i, row in enumerate(self.measurement_rows, start=1):
            h_tot = row.h_feet_total()
            v_tot = row.v_feet_total()
            area_sqft = row.area_sqft()
            area_m2 = area_sqft / 10.7639
            total_sqft += area_sqft
            poly_label = ""
            delta_pct = ""
            if row.polygon_idx is not None and 0 <= row.polygon_idx < len(self.polygons):
                poly = self.polygons[row.polygon_idx]
                poly_area_px2 = polygon_area_image_pts(poly["poly"])
                m2, _ = self.calib.area_real(poly_area_px2)
                poly_sqft = m2 * 10.7639
                poly_label = f"#{row.polygon_idx + 1}"
                if area_sqft > 1e-6:
                    delta_pct = f"{(poly_sqft - area_sqft) / area_sqft * 100:+.2f}%"
                else:
                    delta_pct = "\u2014"
            values = [
                i, row.description, row.shape,
                row.h_subs[0][0], row.h_subs[0][1],
                row.h_subs[1][0], row.h_subs[1][1],
                row.h_subs[2][0], row.h_subs[2][1],
                h_tot,
                row.v_subs[0][0], row.v_subs[0][1],
                row.v_subs[1][0], row.v_subs[1][1],
                row.v_subs[2][0], row.v_subs[2][1],
                v_tot,
                row.factor, area_sqft, area_m2,
                poly_label, delta_pct,
            ]
            for col, v in enumerate(values, start=1):
                c = ws.cell(row=i + 1, column=col, value=v)
                c.border = border
        total_row = len(self.measurement_rows) + 2
        c = ws.cell(row=total_row, column=1, value="Total")
        c.font = bold
        c = ws.cell(row=total_row, column=19, value=total_sqft)
        c.font = bold
        c.number_format = "#,##0.00"
        c = ws.cell(row=total_row, column=20, value=total_sqft / 10.7639)
        c.font = bold
        c.number_format = "#,##0.00"
        widths = [4, 30, 10, 6, 6, 6, 6, 6, 6, 10, 6, 6, 6, 6, 6, 6, 10, 6, 12, 12, 8, 9]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # Sheet 2: Polygons
        wp = wb.create_sheet("Polygons")
        wp.append(["#", "Name", "Colour", "Area (px\u00b2)", "Area (m\u00b2)", "Area (sqft)"])
        for col, _ in enumerate(wp[1], start=1):
            wp[1][col - 1].font = bold
            wp[1][col - 1].fill = header_fill
        for i, p in enumerate(self.polygons, start=1):
            px2 = polygon_area_image_pts(p["poly"])
            m2, _ = self.calib.area_real(px2)
            wp.append([i, p["name"], p["color"], int(px2), m2, m2 * 10.7639])
        for i, w in enumerate([4, 30, 10, 14, 12, 12], start=1):
            wp.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # Sheet 3: Calibration
        wc = wb.create_sheet("Calibration")
        c = self.calib
        wc.append(["Field", "Value"])
        wc["A1"].font = bold; wc["B1"].font = bold
        wc["A1"].fill = header_fill; wc["B1"].fill = header_fill
        wc.append(["V point 1 (px)", c.p1])
        wc.append(["V point 2 (px)", c.p2])
        wc.append(["V real distance", c.v_real])
        wc.append(["H point 1 (px)", c.p3])
        wc.append(["H point 2 (px)", c.p4])
        wc.append(["H real distance", c.h_real])
        wc.append(["Real unit", c.unit])
        wc.append(["Calibrated", "yes" if c.ready else "no"])
        for i, w in enumerate([20, 18], start=1):
            wc.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # Sheet 4: Summary (first tab)
        wsum = wb.create_sheet("Summary", 0)
        wsum.append(["Floor / Plan", os.path.basename(self.image_path or "(no image)")])
        wsum.append(["Generated by", APP_TITLE + "  \u2014  " + (APP_AUTHOR + ", " if APP_AUTHOR else "") + APP_COMPANY])
        wsum.append(["Version", __version__])
        wsum.append(["Total rows", len(self.measurement_rows)])
        wsum.append(["Total area (sqft)", total_sqft])
        wsum.append(["Total area (m\u00b2)", total_sqft / 10.7639])
        for i, w in enumerate([24, 40], start=1):
            wsum.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        try:
            wb.save(path)
            self.status.set(f"Exported measurement sheet to {path}")
        except Exception as e:
            messagebox.showerror("Export", f"Failed to write {path}:\n{e}")

    # ------------------------------------------------------------------
    # Mode change
    # ------------------------------------------------------------------
    def _on_mode_change(self) -> None:
        self.current_poly.clear()
        self.selected_polygon = None
        self.selected_vertex = None
        self._schedule_redraw()

    def _on_input_mode_change(self) -> None:
        if self.input_mode.get() == "pan":
            self.canvas.config(cursor="fleur")
        else:
            self.canvas.config(cursor="crosshair" if self.mode.get() == "polygon" else "")
        self._update_status()

    def _on_poly_list_select(self, _event=None) -> None:
        sel = self.poly_list.curselection()
        if not sel:
            self.selected_polygon = None
            self.selected_vertex = None
        else:
            self.selected_polygon = sel[0]
            self.selected_vertex = None
        self._schedule_redraw()

    # ------------------------------------------------------------------
    # Area formatting
    # ------------------------------------------------------------------
    def format_area(self, area_px2: float) -> str:
        unit = self.unit_var.get()
        if unit == "px²":
            return f"{area_px2:,.0f} px²"
        m2, ft2 = self.calib.area_real(area_px2)
        if unit == "m²":
            return f"{m2:,.2f} m²"
        if unit == "ft²":
            return f"{ft2:,.2f} ft²"
        return ""

    def format_dual_area(self, area_px2: float) -> str:
        m2, ft2 = self.calib.area_real(area_px2)
        return f"{m2:,.2f} m²  /  {ft2:,.2f} ft²"

    # ------------------------------------------------------------------
    # Rendering
    # ------------------------------------------------------------------
    def _schedule_redraw(self) -> None:
        if getattr(self, "_redraw_after", None) is not None:
            self.after_cancel(self._redraw_after)
        self._redraw_after = self.after(15, self.redraw)

    def redraw(self) -> None:
        self._redraw_after = None
        if self.image is None:
            self.canvas.delete("all")
            return

        cw = self.canvas.winfo_width()
        ch = self.canvas.winfo_height()
        if cw < 2 or ch < 2:
            return

        # Build a composite image: base (resized/scaled) + polygon overlays
        # so translucent polygons blend over the floor plan.
        iw, ih = self.image.size
        scaled_w = max(1, int(round(iw * self.zoom)))
        scaled_h = max(1, int(round(ih * self.zoom)))
        if scaled_w <= 0 or scaled_h <= 0:
            return

        # Render the scaled base on a transparent canvas
        rgba = self.image.convert("RGBA")
        scaled = rgba.resize((scaled_w, scaled_h), Image.LANCZOS)
        overlay = Image.new("RGBA", scaled.size, (0, 0, 0, 0))
        drw = ImageDraw.Draw(overlay, "RGBA")

        def to_canvas_poly(poly: Polygon):
            return [(p[0] * self.zoom, p[1] * self.zoom) for p in poly]

        # filled polygons (50% transparent) + outlines
        for p in self.polygons:
            colour = p["color"]
            drw.polygon(to_canvas_poly(p["poly"]), fill=colour + "80", outline=colour + "FF")

        # in-progress polygon
        if self.current_poly and self.mode.get() == "polygon":
            pts = to_canvas_poly(self.current_poly)
            if len(pts) >= 2:
                drw.line(pts, fill="#ffffff", width=2)
            for x, y in pts:
                drw.ellipse((x - 4, y - 4, x + 4, y + 4), fill="#ffffff")
            if self._hover_point:
                hx, hy = self._hover_point[0] * self.zoom, self._hover_point[1] * self.zoom
                pts2 = pts + [(hx, hy)]
                if len(pts2) >= 2:
                    drw.line(pts2, fill="#ffffff", width=1)
            if len(pts) >= 3:
                # preview filled
                drw.polygon(pts, fill="#ffffff40", outline="#ffffff")

        composite = Image.alpha_composite(scaled, overlay)
        self._display_image = composite

        # Convert to Tk image (must keep reference!)
        self.tk_image = ImageTk.PhotoImage(composite)
        self.canvas.delete("all")
        ox, oy = self.offset
        self.canvas.create_image(ox, oy, image=self.tk_image, anchor=tk.NW)

        # Labels (drawn on top in canvas coords; easier crispness)
        if self.show_labels.get():
            self._draw_labels()
        if self.mode.get() == "edit":
            self._draw_edit_handles()

        # Calibration markers
        if self.show_calib.get():
            self._draw_calib()

        self._update_status()

    def _draw_labels(self) -> None:
        for p in self.polygons:
            poly = p["poly"]
            if len(poly) < 3:
                continue
            area_px2 = polygon_area_image_pts(poly)
            text = self.format_area(area_px2)
            # centroid
            cx = sum(pt[0] for pt in poly) / len(poly)
            cy = sum(pt[1] for pt in poly) / len(poly)
            ccx, ccy = self.image_to_canvas(cx, cy)
            self.canvas.create_text(
                ccx, ccy, text=f"{p['name']}\n{text}",
                fill="white", justify=tk.CENTER, font=("Segoe UI", 10, "bold"),
            )

    # ------------------------------------------------------------------
    # Edit mode: hit testing & vertex handles
    # ------------------------------------------------------------------
    def _hit_test_polygon(self, ix, iy, tol_px=8.0):
        """Return (poly_idx, vertex_idx) of the closest vertex within tol_px
        (image-pixel space), or None."""
        best = None
        best_d2 = tol_px * tol_px
        for pi, p in enumerate(self.polygons):
            for vi, (x, y) in enumerate(p["poly"]):
                d2 = (x - ix) ** 2 + (y - iy) ** 2
                if d2 <= best_d2:
                    best = (pi, vi)
                    best_d2 = d2
        return best

    def _point_to_segment_dist(self, px, py, ax, ay, bx, by):
        """Perpendicular distance from (px,py) to segment (ax,ay)-(bx,by)."""
        dx, dy = bx - ax, by - ay
        L2 = dx * dx + dy * dy
        if L2 == 0:
            return math.hypot(px - ax, py - ay)
        t = ((px - ax) * dx + (py - ay) * dy) / L2
        t = max(0.0, min(1.0, t))
        qx, qy = ax + t * dx, ay + t * dy
        return math.hypot(px - qx, py - qy)

    def _hit_test_edge(self, ix, iy, tol_px=8.0):
        """Return (poly_idx, edge_idx) of the closest edge within tol_px."""
        best = None
        best_d = tol_px
        for pi, p in enumerate(self.polygons):
            poly = p["poly"]
            n = len(poly)
            for vi in range(n):
                ax, ay = poly[vi]
                bx, by = poly[(vi + 1) % n]
                d = self._point_to_segment_dist(ix, iy, ax, ay, bx, by)
                if d <= best_d:
                    best = (pi, vi)
                    best_d = d
        return best

    def _draw_edit_handles(self) -> None:
        for pi, p in enumerate(self.polygons):
            is_selected = (pi == self.selected_polygon)
            for vi, (x, y) in enumerate(p["poly"]):
                cx, cy = self.image_to_canvas(x, y)
                if is_selected and vi == self.selected_vertex:
                    r = 8
                    self.canvas.create_oval(cx - r, cy - r, cx + r, cy + r,
                                            fill="white", outline="black", width=2)
                    self.canvas.create_text(
                        cx + 12, cy - 12, anchor=tk.NW,
                        text=f"({x:.1f}, {y:.1f})",
                        fill="yellow", font=("Consolas", 9, "bold"))
                else:
                    r = 5
                    fill = "white" if is_selected else "#222"
                    self.canvas.create_oval(cx - r, cy - r, cx + r, cy + r,
                                            fill=fill, outline=p["color"], width=2)
            if self._hover_edge is not None and self._hover_edge[0] == pi:
                _, ei = self._hover_edge
                ax, ay = p["poly"][ei]
                bx, by = p["poly"][(ei + 1) % len(p["poly"])]
                mx, my = (ax + bx) / 2, (ay + by) / 2
                cmx, cmy = self.image_to_canvas(mx, my)
                r = 4
                self.canvas.create_oval(cmx - r, cmy - r, cmx + r, cmy + r,
                                        fill="yellow", outline="black", width=1)

    def _draw_calib(self) -> None:
        c = self.calib
        for i, (pt, colour) in enumerate([
            (c.p1, "#ff5252"), (c.p2, "#ff5252"),
            (c.p3, "#42a5f5"), (c.p4, "#42a5f5"),
        ]):
            if pt is None:
                continue
            cx, cy = self.image_to_canvas(*pt)
            r = 6
            self.canvas.create_oval(cx - r, cy - r, cx + r, cy + r,
                                    fill=colour, outline="white", width=2)
        # lines + distance labels
        if c.p1 and c.p2:
            x1, y1 = self.image_to_canvas(*c.p1)
            x2, y2 = self.image_to_canvas(*c.p2)
            self.canvas.create_line(x1, y1, x2, y2, fill="#ff5252", dash=(4, 3), width=2)
            mx, my = (x1 + x2) / 2, (y1 + y2) / 2
            label = f"V: {c.v_real:g} {c.unit}" if c.v_real else "V: ?"
            self.canvas.create_text(mx, my - 10, text=label, fill="#ff5252",
                                    font=("Segoe UI", 10, "bold"))
        if c.p3 and c.p4:
            x1, y1 = self.image_to_canvas(*c.p3)
            x2, y2 = self.image_to_canvas(*c.p4)
            self.canvas.create_line(x1, y1, x2, y2, fill="#42a5f5", dash=(4, 3), width=2)
            mx, my = (x1 + x2) / 2, (y1 + y2) / 2
            label = f"H: {c.h_real:g} {c.unit}" if c.h_real else "H: ?"
            self.canvas.create_text(mx, my - 10, text=label, fill="#42a5f5",
                                    font=("Segoe UI", 10, "bold"))

    def _update_status(self) -> None:
        if self.image is None:
            self.status.set("Open an image to start.")
            return
        iw, ih = self.image.size
        n = len(self.polygons)
        cur = len(self.current_poly)
        mode = self.mode.get()
        inm = self.input_mode.get().upper()
        bits = [
            f"{APP_TITLE} — {APP_AUTHOR}, {APP_COMPANY}",
            f"image: {iw}×{ih} px",
            f"mode: {mode}",
            f"input: {inm}",
            f"zoom: {self.zoom*100:.0f}%",
            f"polygons: {n}",
            f"current pts: {cur}",
        ]
        self.status.set("   |   ".join(bits))

    # ------------------------------------------------------------------
    # About / version
    # ------------------------------------------------------------------
    def _show_about(self) -> None:
        ver = globals().get("__version__", "dev")
        build = globals().get("__build__", "local")
        built = globals().get("__built_at__", "")
        messagebox.showinfo(
            f"About — {APP_TITLE}",
            f"{APP_TITLE}\n"
            f"\n"
            f"Created by:  {APP_AUTHOR}\n"
            f"Company:     {APP_COMPANY}\n"
            f"\n"
            f"Version: {ver}\n"
            f"Build:   {build}\n"
            f"Built:   {built or '(local source)'}\n"
            f"\n"
            f"MIT License\n"
            f"https://github.com/choumakdou/floor-plan-area-meter",
        )

    # ------------------------------------------------------------------
    # Project save / load
    # ------------------------------------------------------------------
    def save_project(self) -> None:
        if self.image is None:
            messagebox.showinfo("Save", "Open an image first.")
            return
        path = filedialog.asksaveasfilename(
            title="Save project",
            defaultextension=".fpm.json",
            filetypes=[("Floor Plan Meter project", "*.fpm.json")],
        )
        if not path:
            return
        data = {
            "image": self.image_path,
            "calib": self.calib.to_dict(),
            "polygons": [
                {"poly": p["poly"], "color": p["color"], "name": p["name"]}
                for p in self.polygons
            ],
            "measurement_rows": [r.to_dict() for r in self.measurement_rows],
        }
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
        self.status.set(f"Saved project → {path}")

    def load_project(self) -> None:
        path = filedialog.askopenfilename(
            title="Load project",
            filetypes=[("Floor Plan Meter project", "*.fpm.json"),
                       ("JSON", "*.json")],
        )
        if not path:
            return
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        img_path = data.get("image")
        if not img_path or not os.path.exists(img_path):
            messagebox.showwarning("Load", "Image file from project not found:\n" + str(img_path))
            return
        self.image = Image.open(img_path); self.image.load()
        self.image_path = img_path
        self.calib = Calibration(); self.calib.from_dict(data.get("calib", {}))
        self.polygons = data.get("polygons", [])
        self.poly_list.delete(0, tk.END)
        for p in self.polygons:
            self.poly_list.insert(tk.END, p.get("name", "?"))
        self.current_poly.clear()
        # Restore measurement rows (v2.1).  Missing key -> empty list.
        self.measurement_rows = [MeasurementRow.from_dict(d) for d in data.get("measurement_rows", [])]
        self._refresh_measurement_sheet()
        self._refresh_calib_entries()
        self.fit_to_window()
        self._update_status()


# ---------------------------------------------------------------------------
def main() -> None:
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
